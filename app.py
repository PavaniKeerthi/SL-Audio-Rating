import streamlit as st
from google.oauth2 import service_account
from google.cloud import storage
import gspread
import pandas as pd
import librosa
import numpy as np
import io
import tempfile
import os
from fpdf import FPDF
import openpyxl
import textwrap

# --- Page and App Configuration ---
st.set_page_config(layout="wide", page_title="Audio Quality Rater")

# --- Authentication and Client Functions ---
@st.cache_resource
def get_gcs_client():
    """Initializes and returns a Google Cloud Storage client using Streamlit secrets."""
    try:
        creds_info = st.secrets["gcp_service_account"]
        credentials = service_account.Credentials.from_service_account_info(creds_info)
        return storage.Client(credentials=credentials)
    except Exception as e:
        st.error(f"GCS Auth Error: {e}")
        return None

@st.cache_resource
def get_gspread_client():
    """Initializes and returns a gspread client for Google Sheets."""
    try:
        creds_info = st.secrets["gcp_service_account"]
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        credentials = service_account.Credentials.from_service_account_info(
            creds_info, scopes=scopes
        )
        return gspread.authorize(credentials)
    except Exception as e:
        st.error(f"Google Sheets Auth Error: {e}")
        return None

# --- Audio Analysis and GCS Functions ---
def download_audio_from_gcs(gcs_url, client):
    """Downloads an audio file from a GCS URL into an in-memory stream."""
    try:
        clean_url = gcs_url.strip()
        if not clean_url or not isinstance(clean_url, str) or not clean_url.startswith("https://storage.cloud.google.com/"):
            return None
        path_parts = clean_url.replace("https://storage.cloud.google.com/", "").split('/')
        bucket_name = path_parts[0]
        blob_name = "/".join(path_parts[1:])
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        if not blob.exists():
            st.warning(f"File not found in GCS: {blob_name}")
            return None
        audio_data = blob.download_as_bytes()
        return io.BytesIO(audio_data)
    except Exception as e:
        st.warning(f"Could not download {gcs_url}: {e}")
        return None

def analyze_audio_quality(file_path, url_for_logging):
    """Analyzes audio quality and returns both display score and numerical score."""
    try:
        if url_for_logging.strip().endswith('.txt'):
            return None
        y, sr = librosa.load(file_path, sr=None)
        rms = librosa.feature.rms(y=y, frame_length=2048, hop_length=512)[0]
        if len(rms) > 0 and np.max(rms) > 0:
            signal_power = np.mean(rms**2)
            noise_power = np.min(rms*2) if np.min(rms*2) > 0 else 1e-10
            snr = 10 * np.log10(signal_power / noise_power)
        else:
            snr = 0
        clipping_percentage = np.mean(np.abs(y) >= 0.99) * 100
        total_frames = len(rms)
        silence_percentage = (np.sum(rms < 0.01) / total_frames) * 100 if total_frames > 0 else 100
        score = 1.0
        if snr > 15: score += 4.0
        elif snr > 12: score += 3.0
        elif snr > 8: score += 1.5
        if clipping_percentage < 1: score += 3.0
        elif clipping_percentage < 5: score += 1.5
        elif clipping_percentage < 10: score += 0.5
        if silence_percentage < 15: score += 2.0
        elif silence_percentage < 30: score += 1.0
        elif silence_percentage < 45: score += 0.5
        final_score = min(score, 10.0)
        return {"display_score": f"{final_score:.1f} / 10", "numerical_score": final_score}
    except Exception as e:
        st.warning(f"Could not analyze audio from URL {url_for_logging}. Reason: {e}")
        return None

# --- Report Generation Functions ---
def dataframe_to_excel_bytes(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Audio_Quality_Report')
    return output.getvalue()

class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'Audio Quality Analysis Report', 0, 1, 'C')
        self.ln(5)
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

# --- PDF Table Headings Improvement ---
def create_pdf_report(df):
    pdf_col_map = {
        'Session Key': 'Session Key',
        'Average Audio Score': 'Avg Score',
        'Response1': 'Prompt1',
        'AQGResponse_P1_1': 'AQG1',
        'AQGResponse_P1_2': 'AQG2',
        'Response2': 'Prompt2',
        'AQGResponse_P2_1': 'AQG1',
        'AQGResponse_P2_2': 'AQG2',
        'Response3': 'Prompt3',
        'AQGResponse_P3_1': 'AQG1',
        'AQGResponse_P3_2': 'AQG2',
        'Response4': 'Prompt4',
        'AQGResponse_P4_1': 'AQG1',
        'AQGResponse_P4_2': 'AQG2',
        'QResponse1': 'Question1',
        'QResponse2': 'Question2',
        'QResponse3': 'Question3',
        'QResponse4': 'Question4',
        'QResponse5': 'Question5',
    }
    pdf_df = df.rename(columns=pdf_col_map)
    pdf = PDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 10, 'Individual Audio Analysis Results', 0, 1, 'L')
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Arial', 'B', 8)
    headers = list(pdf_df.columns)
    # Custom widths: Session Key (28), Avg Score (16), Prompt (11), AQG (7), Question (20)
    col_widths = []
    for h in headers:
        if h == 'Session Key':
            col_widths.append(22)
        elif h == 'Avg Score':
            col_widths.append(16)
        elif h.startswith('Prompt'):
            col_widths.append(11)
        elif h.startswith('AQG'):
            col_widths.append(9)
        elif h.startswith('Question'):
            col_widths.append(20)
        else:
            col_widths.append(11)
    # Scale to fit A4 landscape (277mm)
    total_width = sum(col_widths)
    scale = 277 / total_width
    col_widths = [w * scale for w in col_widths]
    # Print header
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 8, header, border=1, align='C', fill=True)
    pdf.ln(8)
    # Data rows
    pdf.set_font('Arial', '', 7)
    pdf.set_text_color(0,0,0)
    fill = False
    for idx, (_, row) in enumerate(pdf_df.iterrows()):
        if fill:
            pdf.set_fill_color(240, 245, 255)
        else:
            pdf.set_fill_color(255,255,255)
        for i, item in enumerate(row):
            pdf.cell(col_widths[i], 7, str(item), border=1, align='C', fill=True)
        pdf.ln(7)
        fill = not fill
    # Methodology and Recommendations
    pdf.add_page(orientation='P')
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'Rating Methodology', 0, 1, 'L')
    pdf.set_font('Arial', '', 10)
    pdf.multi_cell(0, 5,
        "The overall audio quality score (from 1 to 10) is calculated based on several key technical metrics. "
        "A higher score indicates better audio quality."
    )
    pdf.ln(5)
    methodology = {
        "Signal-to-Noise Ratio (SNR):": "Measures the clarity of the speech compared to background noise. A higher SNR is better.",
        "Clipping Percentage:": "Indicates distortion caused when the audio signal is too loud. Less than 1% is ideal.",
        "Silence Percentage:": "Represents the amount of dead air or excessive pauses. Too much silence can negatively impact listener engagement."
    }
    for title, desc in methodology.items():
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 6, title, 0, 1, 'L')
        pdf.set_font('Arial', '', 10)
        pdf.multi_cell(0, 5, desc)
        pdf.ln(2)
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'Recommendations for Better Ratings', 0, 1, 'L')
    pdf.set_font('Arial', '', 10)
    recommendations = {
        "To Improve SNR:": "Use a high-quality microphone, record in a quiet environment, and position the mic correctly.",
        "To Avoid Clipping:": "Monitor your recording levels. Keep the input level out of the 'red' zone (ideally peaking around -6dB).",
        "To Reduce Excessive Silence:": "Plan your content to ensure a steady flow. Use audio editing software to trim long pauses.",
        "General Best Practices:": "Record in a high-resolution format (e.g., WAV). Ensure consistent microphone distance and speaking volume."
    }
    for title, desc in recommendations.items():
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 6, title, 0, 1, 'L')
        pdf.set_font('Arial', '', 10)
        pdf.multi_cell(0, 5, desc)
        pdf.ln(2)
    return bytes(pdf.output(dest='S'))

# --- Main Application UI ---
st.title("🎧 Audio & Transcript Quality Dashboard")

# --- Top right horizontal evaluation buttons (absolute top right) ---
btn_spacer, btn_col1, btn_col2 = st.columns([6,1,1], gap="small")
with btn_col1:
    audio_eval_btn = st.button("🎵 Audio Quality Evaluation", key="audio_eval_btn", use_container_width=True)
with btn_col2:
    transcript_eval_btn = st.button("📝 Transcript Quality Evaluation", key="transcript_eval_btn", use_container_width=True)

# --- Prominent Dashboard Card (only the word 'Dashboard') ---
st.markdown("""
<div style='background: linear-gradient(90deg, #dbeafe 0%, #f0fdfa 100%); padding: 2em 2em 1em 2em; border-radius: 18px; box-shadow: 0 4px 24px #0002; margin-bottom:1em;'>
    <h2 style='text-align:center; color:#1e293b; letter-spacing:1px; margin-bottom:0.5em;'>Dashboard</h2>
</div>
""", unsafe_allow_html=True)

# --- Instructions (outside dashboard box) ---
st.markdown("""
<div style='background: #f8fafc; border-radius: 12px; padding: 1.2em 2em; margin-bottom:1.5em; border: 1px solid #e0e7ef;'>
    <ol style='font-size:1.1em; color:#334155;'>
        <li>Share your Google Sheet with the service account email.</li>
        <li>Sheet must have a tab named <b>URLs</b>.</li>
        <li>Tab must have columns <b>Session Key</b> and <b>Status</b>.</li>
        <li>Paste the Google Sheet URL, select status, and use the buttons above.</li>
    </ol>
</div>
""", unsafe_allow_html=True)

# --- Inputs ---
sheet_url = st.text_input("🔗 Enter your Google Sheet URL:")
status_options = ["All", "Responded", "Not Responded"]
selected_status = st.selectbox("📋 Select the status to analyze:", options=status_options)

gspread_client = get_gspread_client()
gcs_client = get_gcs_client()

if not gspread_client or not gcs_client:
    st.stop()

if 'results_df' not in st.session_state:
    st.session_state.results_df = None
if 'evaluation_done' not in st.session_state:
    st.session_state.evaluation_done = False

AUDIO_COLUMN_NAMES = [
    "Response1", "AQGResponse_P1_1", "AQGResponse_P1_2", "Response2", "AQGResponse_P2_1", "AQGResponse_P2_2",
    "Response3", "AQGResponse_P3_1", "AQGResponse_P3_2", "Response4", "AQGResponse_P4_1", "AQGResponse_P4_2",
    "QResponse1", "QResponse2", "QResponse3", "QResponse4", "QResponse5"
]

gen_btn = st.button("🚀 Generate", key="generate_btn", use_container_width=True)

if sheet_url and gen_btn:
    try:
        spreadsheet = gspread_client.open_by_url(sheet_url)
        SHEET_NAME = "URLs"
        SESSION_KEY_COL = "Session Key"
        STATUS_COL = "Status"
        try:
            worksheet = spreadsheet.worksheet(SHEET_NAME)
        except gspread.exceptions.WorksheetNotFound:
            st.error(f"❌ *Error:* A worksheet named *'{SHEET_NAME}'* was not found.")
            st.stop()
        df = pd.DataFrame(worksheet.get_all_records(head=1, default_blank=''))
        if df.empty:
            st.warning(f"The sheet '{SHEET_NAME}' is loaded but appears to be empty.")
            st.stop()
        df = df.astype(str)
        if SESSION_KEY_COL not in df.columns or STATUS_COL not in df.columns:
            st.error(f"❌ *Error:* Required columns '{SESSION_KEY_COL}' and/or '{STATUS_COL}' not found.")
            st.stop()
        if selected_status == "All":
            df_to_process = df.copy()
        else:
            df_to_process = df[df[STATUS_COL] == selected_status].copy()
        if df_to_process.empty:
            st.warning(f"No rows found with the status '{selected_status}'.")
            st.session_state.results_df = None
            st.session_state.evaluation_done = False
        else:
            all_results = []
            total_rows = len(df_to_process)
            progress_bar = st.progress(0, text=f"Starting analysis on {total_rows} row(s)...")
            for i, (index, row) in enumerate(df_to_process.iterrows()):
                session_key = row[SESSION_KEY_COL]
                progress_text = f"Analyzing row for Session Key: {session_key} ({i + 1}/{total_rows})"
                progress_bar.progress((i + 1) / total_rows, text=progress_text)
                current_row_result = {"Session Key": session_key}
                row_scores = []
                for col_name in AUDIO_COLUMN_NAMES:
                    if col_name in row and row[col_name].strip().startswith("https://storage.cloud.google.com/"):
                        cell_value = row[col_name]
                        audio_stream = download_audio_from_gcs(cell_value, gcs_client)
                        if audio_stream:
                            file_suffix = os.path.splitext(cell_value)[1] or ".tmp"
                            with tempfile.NamedTemporaryFile(delete=False, suffix=file_suffix) as temp_file:
                                temp_file.write(audio_stream.read())
                                temp_file_path = temp_file.name
                            try:
                                rating = analyze_audio_quality(temp_file_path, cell_value)
                                if rating:
                                    current_row_result[col_name] = rating["display_score"]
                                    row_scores.append(rating["numerical_score"])
                                else:
                                    current_row_result[col_name] = "Analysis Error"
                            finally:
                                if os.path.exists(temp_file_path):
                                    os.remove(temp_file_path)
                        else:
                            current_row_result[col_name] = "Download Fail"
                    else:
                        current_row_result[col_name] = "N/A"
                if row_scores:
                    average_score = sum(row_scores) / len(row_scores)
                    current_row_result["Average Audio Score"] = f"{average_score:.1f} / 10"
                else:
                    current_row_result["Average Audio Score"] = "N/A"
                all_results.append(current_row_result)
            progress_bar.empty()
            if all_results:
                results_df = pd.DataFrame(all_results)
                final_column_order = ["Session Key", "Average Audio Score"] + AUDIO_COLUMN_NAMES
                results_df = results_df[final_column_order]
                st.session_state.results_df = results_df
                st.session_state.evaluation_done = True
                st.success("Evaluation complete! Click 'Audio Quality Evaluation' to view and download your results.")
            else:
                st.session_state.results_df = None
                st.session_state.evaluation_done = False
    except gspread.exceptions.SpreadsheetNotFound:
        st.error("❌ *Error:* Spreadsheet not found. Check the URL and sharing permissions.")
    except Exception as e:
        st.error(f"An unexpected error occurred: {e}")

# --- Show results only when Audio Quality Evaluation is pressed ---
if audio_eval_btn:
    if st.session_state.get('results_df') is not None and st.session_state.get('evaluation_done'):
        st.subheader("🎵 Audio Quality Analysis Results")
        st.dataframe(st.session_state.results_df)
        st.subheader("⬇️ Download Reports")
        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            excel_bytes = dataframe_to_excel_bytes(st.session_state.results_df)
            st.download_button(label="📥 Download Excel Report", data=excel_bytes, file_name="audio_quality_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with dl_col2:
            pdf_bytes = create_pdf_report(st.session_state.results_df)
            st.download_button(label="📄 Download PDF Report", data=pdf_bytes, file_name="audio_quality_report.pdf", mime="application/pdf")
    else:
        st.info("Please run 'Generate' first to perform the evaluation.")

# --- Placeholder for Transcript Quality Evaluation ---
if transcript_eval_btn:
    st.subheader("📝 Transcript Quality Evaluation")
    st.info("Transcript quality evaluation feature coming soon!")
